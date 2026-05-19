import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter
import os

# ── Colors ─────────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1B3A6B")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL     = PatternFill("solid", fgColor="D6E4F0")
GREEN_FILL   = PatternFill("solid", fgColor="C6EFCE")
RED_FILL     = PatternFill("solid", fgColor="FFC7CE")
TITLE_FONT   = Font(bold=True, size=14, color="1B3A6B")
SUBTITLE_FONT = Font(bold=True, size=11, color="2E86AB")

thin  = Side(style='thin', color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header_row(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER


def style_data_rows(ws, start_row, end_row, ncols):
    for r in range(start_row, end_row + 1):
        fill = ALT_FILL if r % 2 == 0 else PatternFill()
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center')


def load_data():
    return pd.read_csv('exports/financials.csv')


# ── Sheet 1: Raw Data ──────────────────────────────────────────────────────────
def sheet_raw_data(wb, df):
    ws = wb.create_sheet("Raw Data")
    headers = list(df.columns)
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    for _, row in df.iterrows():
        ws.append(list(row))
    style_data_rows(ws, 2, len(df) + 1, len(headers))
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.freeze_panes = "A2"
    print("Sheet 1: Raw Data done")


# ── Sheet 2: P&L Statement ─────────────────────────────────────────────────────
def sheet_pnl(wb, df):
    ws = wb.create_sheet("P&L Statement")
    ws['A1'] = "Profit & Loss Statement"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:E1')

    for year in [2022, 2023, 2024]:
        yr_df = df[df['year'] == year]
        row_start = ws.max_row + 2

        ws.cell(row=row_start, column=1, value=f"Year {year}").font = SUBTITLE_FONT
        row_start += 1

        headers = ['Month', 'Revenue (AED)', 'Expense (AED)', 'Profit (AED)', 'Margin %']
        for col, h in enumerate(headers, 1):
            ws.cell(row=row_start, column=col, value=h)
        style_header_row(ws, row_start, len(headers))
        row_start += 1

        monthly = yr_df.groupby('month_name').agg(
            Revenue=('revenue', 'sum'),
            Expense=('expense', 'sum'),
            Profit=('profit', 'sum')
        ).reset_index()
        monthly['Margin'] = (monthly['Profit'] / monthly['Revenue'] * 100).round(2)

        month_order = ['January','February','March','April','May','June',
                       'July','August','September','October','November','December']
        monthly['month_name'] = pd.Categorical(monthly['month_name'], categories=month_order, ordered=True)
        monthly = monthly.sort_values('month_name')

        for i, (_, row) in enumerate(monthly.iterrows()):
            r = row_start + i
            ws.cell(row=r, column=1, value=row['month_name'])
            ws.cell(row=r, column=2, value=round(row['Revenue'], 2))
            ws.cell(row=r, column=3, value=round(row['Expense'], 2))
            ws.cell(row=r, column=4, value=round(row['Profit'], 2))
            ws.cell(row=r, column=5, value=round(row['Margin'], 2))

            fill = ALT_FILL if i % 2 == 0 else PatternFill()
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = fill
                ws.cell(row=r, column=c).border = BORDER
                ws.cell(row=r, column=c).alignment = Alignment(horizontal='center')

            # Color profit cell
            profit_cell = ws.cell(row=r, column=4)
            if row['Profit'] < 0:
                profit_cell.fill = RED_FILL
            else:
                profit_cell.fill = GREEN_FILL

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 22
    print("Sheet 2: P&L Statement done")


# ── Sheet 3: Budget vs Actual ──────────────────────────────────────────────────
def sheet_budget_vs_actual(wb, df):
    ws = wb.create_sheet("Budget vs Actual")
    ws['A1'] = "Budget vs Actual Analysis"
    ws['A1'].font = TITLE_FONT

    bva = df.groupby('department').agg(
        Actual_Revenue=('revenue', 'sum'),
        Budget_Revenue=('budget_revenue', 'sum'),
        Actual_Expense=('expense', 'sum'),
        Budget_Expense=('budget_expense', 'sum'),
    ).reset_index()

    bva['Revenue_Variance'] = (bva['Actual_Revenue'] - bva['Budget_Revenue']).round(2)
    bva['Expense_Variance'] = (bva['Actual_Expense'] - bva['Budget_Expense']).round(2)
    bva['Revenue_Variance_%'] = ((bva['Revenue_Variance'] / bva['Budget_Revenue']) * 100).round(2)

    headers = ['Department', 'Actual Revenue', 'Budget Revenue', 'Revenue Variance',
               'Variance %', 'Actual Expense', 'Budget Expense', 'Expense Variance']
    ws.append([])
    ws.append(headers)
    style_header_row(ws, 3, len(headers))

    for i, (_, row) in enumerate(bva.iterrows()):
        r = 4 + i
        ws.append([
            row['department'],
            round(row['Actual_Revenue'], 2),
            round(row['Budget_Revenue'], 2),
            round(row['Revenue_Variance'], 2),
            round(row['Revenue_Variance_%'], 2),
            round(row['Actual_Expense'], 2),
            round(row['Budget_Expense'], 2),
            round(row['Expense_Variance'], 2),
        ])
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = fill
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = Alignment(horizontal='center')

        var_cell = ws.cell(row=r, column=4)
        var_cell.fill = GREEN_FILL if row['Revenue_Variance'] >= 0 else RED_FILL

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Bar chart
    chart = BarChart()
    chart.title = "Actual vs Budget Revenue by Department"
    chart.y_axis.title = "Amount (AED)"
    data = Reference(ws, min_col=2, max_col=3, min_row=3, max_row=3 + len(bva))
    cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(bva))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 22
    chart.height = 12
    ws.add_chart(chart, "J3")
    print("Sheet 3: Budget vs Actual done")


# ── Sheet 4: Monthly Trends ────────────────────────────────────────────────────
def sheet_monthly_trends(wb, df):
    ws = wb.create_sheet("Monthly Trends")
    ws['A1'] = "Monthly Revenue & Expense Trends"
    ws['A1'].font = TITLE_FONT

    monthly = df.groupby(['year', 'month']).agg(
        Revenue=('revenue', 'sum'),
        Expense=('expense', 'sum'),
        Profit=('profit', 'sum')
    ).reset_index().sort_values(['year', 'month'])
    monthly['Period'] = monthly['year'].astype(str) + '-' + monthly['month'].astype(str).str.zfill(2)

    headers = ['Period', 'Revenue (AED)', 'Expense (AED)', 'Profit (AED)']
    ws.append([])
    ws.append(headers)
    style_header_row(ws, 3, len(headers))

    for i, (_, row) in enumerate(monthly.iterrows()):
        r = 4 + i
        ws.append([row['Period'], round(row['Revenue'], 2),
                   round(row['Expense'], 2), round(row['Profit'], 2)])
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = fill
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = Alignment(horizontal='center')

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Line chart
    chart = LineChart()
    chart.title = "Revenue vs Expense Trend"
    chart.y_axis.title = "Amount (AED)"
    data = Reference(ws, min_col=2, max_col=3, min_row=3, max_row=3 + len(monthly))
    chart.add_data(data, titles_from_data=True)
    chart.width = 22
    chart.height = 12
    ws.add_chart(chart, "F3")
    print("Sheet 4: Monthly Trends done")


# ── Sheet 5: Department Cost Breakdown ────────────────────────────────────────
def sheet_dept_costs(wb, df):
    ws = wb.create_sheet("Dept Cost Breakdown")
    ws['A1'] = "Department Cost Breakdown"
    ws['A1'].font = TITLE_FONT

    dept = df.groupby('department').agg(
        Total_Revenue=('revenue', 'sum'),
        Total_Expense=('expense', 'sum'),
        Total_Profit=('profit', 'sum'),
        Avg_Margin=('profit_margin', 'mean')
    ).reset_index()
    dept = dept.round(2)

    headers = ['Department', 'Total Revenue', 'Total Expense', 'Total Profit', 'Avg Margin %']
    ws.append([])
    ws.append(headers)
    style_header_row(ws, 3, len(headers))

    for i, (_, row) in enumerate(dept.iterrows()):
        r = 4 + i
        ws.append(list(row))
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = fill
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = Alignment(horizontal='center')

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 22

    # Pie chart
    chart = PieChart()
    chart.title = "Revenue by Department"
    data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(dept))
    cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(dept))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 18
    chart.height = 12
    ws.add_chart(chart, "G3")
    print("Sheet 5: Dept Cost Breakdown done")


# ── Sheet 6: Executive Dashboard ──────────────────────────────────────────────
def sheet_dashboard(wb, df):
    ws = wb.create_sheet("Executive Dashboard")
    ws.sheet_properties.tabColor = "1B3A6B"

    ws['A1'] = "RevenueRadar — Executive Financial Dashboard"
    ws['A1'].font = Font(bold=True, size=16, color="1B3A6B")
    ws.merge_cells('A1:D1')

    total_revenue  = round(df['revenue'].sum(), 2)
    total_expense  = round(df['expense'].sum(), 2)
    total_profit   = round(df['profit'].sum(), 2)
    avg_margin     = round(df['profit_margin'].mean(), 2)
    best_dept      = df.groupby('department')['profit'].sum().idxmax()
    worst_dept     = df.groupby('department')['profit'].sum().idxmin()

    kpis = [
        ("Total Revenue (AED)", f"{total_revenue:,.2f}"),
        ("Total Expense (AED)", f"{total_expense:,.2f}"),
        ("Total Profit (AED)", f"{total_profit:,.2f}"),
        ("Avg Profit Margin %", f"{avg_margin}%"),
        ("Most Profitable Dept", best_dept),
        ("Least Profitable Dept", worst_dept),
    ]

    ws['A3'] = "KPI"
    ws['B3'] = "Value"
    style_header_row(ws, 3, 2)

    for i, (kpi, val) in enumerate(kpis, start=4):
        ws.cell(row=i, column=1, value=kpi).font = Font(bold=True)
        ws.cell(row=i, column=2, value=val)
        ws.cell(row=i, column=1).border = BORDER
        ws.cell(row=i, column=2).border = BORDER

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 25
    print("Sheet 6: Executive Dashboard done")


# ── Main ───────────────────────────────────────────────────────────────────────
def generate_excel_report():
    df = load_data()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sheet_raw_data(wb, df)
    sheet_pnl(wb, df)
    sheet_budget_vs_actual(wb, df)
    sheet_monthly_trends(wb, df)
    sheet_dept_costs(wb, df)
    sheet_dashboard(wb, df)

    os.makedirs('exports/excel', exist_ok=True)
    path = 'exports/excel/RevenueRadar_Financial_Report.xlsx'
    wb.save(path)
    print(f"\nExcel report saved to {path}")


if __name__ == "__main__":
    generate_excel_report()